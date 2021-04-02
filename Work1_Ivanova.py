from openpyxl import load_workbook
import geojson
import random
from pymongo import MongoClient


def opener():
    exl = load_workbook('data.xlsx')
    list = exl['Лист1']
    id = []
    x = []
    y = []
    for i in range(1,129):
        val = list.cell(row=i+1, column=1).value
        id.append(val)
    for i in range (1, 129):
        val = list.cell(row=i+1, column=2).value
        x.append(float(val))
    for i in range (1, 129):
        val = list.cell(row=i+1, column=3).value
        y.append(float(val))
    return id, x, y

def sample(property, type):
    client = MongoClient('localhost', 27017)
    db = client['MyDB']
    collection_currency = db['MyColl']
    # Создаём пустые фичи
    new_features = []
    for i in collection_currency.find({property: type}):
        new_features.append(geojson.Feature(geometry=i['geometry'], properties=i['properties']))
    new_feature_collection = geojson.FeatureCollection(features=new_features)
    # Генерируем файл выборки с именем Sample + random
    with open('Sample'+str(random.choice(range(1000,10000))), 'w') as r:
        geojson.dump(new_feature_collection, r)

id, x, y = opener()
coordinates = zip(y, x)
coordinates = list(coordinates)

features = []
for i in coordinates:
    width = [6, 10, 12]
    speed = [60, 80, 100, 120]
    type_of_surface = ["ground","city", "highway", "expressway"]
    lanes = [2, 4]
    features.append(geojson.Feature(geometry=geojson.Point(i), properties={"width":random.choice(width),
                                                                           "speed":random.choice(speed),
                                                                           "type of surface":random.choice(type_of_surface),
                                                                           "lanes":random.choice(lanes)}))
collection = geojson.FeatureCollection(features=features)

with open('Points.geojson', 'w') as file:
    geojson.dump(collection, file)
with open('Points_for_DB.json', 'w') as file:
    geojson.dump(features, file)

client = MongoClient('localhost', 27017)
db = client['MyDB']
collection_currency = db['MyColl']

# Подключение 1 раз, потом шарп
# with open('Points_for_DB.json') as f:
    # file_data = geojson.load(f)
# collection_currency.insert_many(file_data)

sample('properties.width', 10)
sample('properties.speed', 80)
sample('properties.type of surface', 'city')
sample('properties.width', 12)
sample('properties.type of surface', 'ground')
sample('properties.lanes',4)
sample('properties.type of surface', 'highway')
sample('properties.lanes', 2)
sample('properties.speed', 60)
sample('properties.speed', 120)

